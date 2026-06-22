from datetime import datetime
from functools import wraps
import os
import csv
import json
from textwrap import wrap

from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from database import fetch_all, fetch_one, execute, execute_scalar, get_connection

app = Flask(__name__)
os.makedirs("facturas_pdf", exist_ok=True)
os.makedirs("ordenes_pdf", exist_ok=True)
app.secret_key = os.getenv("SECRET_KEY", "palmar-clave-secreta-cambiar")


# =========================
# SEGURIDAD
# =========================
def login_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapper


def admin_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if session.get("rol") != "Administrador":
            flash("No tienes permiso para entrar a este módulo", "warning")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapper


@app.context_processor
def inject_user():
    return {"usuario_actual": session.get("usuario"), "rol_actual": session.get("rol")}


# =========================
# LOGIN
# =========================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        password = request.form.get("password", "").strip()
        user = fetch_one("SELECT id, usuario, rol FROM usuarios WHERE usuario=? AND pass=?", (usuario, password))
        if user:
            session["usuario"] = user["usuario"]
            session["rol"] = user["rol"]
            return redirect(url_for("dashboard"))
        flash("Usuario o contraseña incorrectos", "danger")
    return render_template("login.html")


@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        password = request.form.get("password", "").strip()
        confirmar = request.form.get("confirmar", "").strip()
        clave = request.form.get("clave", "").strip()
        rol = "Cliente"

        if not usuario or not password or not confirmar:
            flash("Por favor, completa todos los campos", "danger")
            return render_template("registro.html")
        if len(password) < 6:
            flash("La contraseña debe tener al menos 6 caracteres", "danger")
            return render_template("registro.html")
        if password != confirmar:
            flash("Las contraseñas no coinciden", "danger")
            return render_template("registro.html")
        if clave != "1234":
            flash("La clave de registro es incorrecta", "danger")
            return render_template("registro.html")
        #if rol not in ("Cliente", "Vendedor"):
            #rol = "Cliente"

        existe = fetch_one("SELECT id FROM usuarios WHERE usuario=?", (usuario,))
        if existe:
            flash("El usuario ya existe", "danger")
            return render_template("registro.html")

        execute("INSERT INTO usuarios(usuario, pass, rol) VALUES (?, ?, ?)", (usuario, password, rol))
        flash("Usuario registrado correctamente. Ahora puedes iniciar sesión", "success")
        return redirect(url_for("login"))
    return render_template("registro.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# =========================
# DASHBOARD
# =========================
@app.route("/dashboard")
@login_required
def dashboard():
    ventas_7 = fetch_all("""
        SELECT 
            TO_CHAR((fecha AT TIME ZONE 'UTC' AT TIME ZONE 'America/Managua')::date, 'YYYY-MM-DD') AS dia,
            COUNT(*) AS cantidad,
            COALESCE(SUM(total), 0) AS total
        FROM ventas
        WHERE (fecha AT TIME ZONE 'UTC' AT TIME ZONE 'America/Managua')::date >=
            ((CURRENT_TIMESTAMP AT TIME ZONE 'America/Managua')::date - INTERVAL '6 days')
        GROUP BY (fecha AT TIME ZONE 'UTC' AT TIME ZONE 'America/Managua')::date
        ORDER BY (fecha AT TIME ZONE 'UTC' AT TIME ZONE 'America/Managua')::date
    """)

    mas_vendidos = fetch_all("""
        SELECT p.nombre, SUM(d.cantidad) AS unidades
        FROM detalle_venta d
        INNER JOIN productos p ON d.producto_id = p.id
        GROUP BY p.nombre
        ORDER BY SUM(d.cantidad) DESC
        LIMIT 4
    """)

    actividad = fetch_all("""
        SELECT *
        FROM (
            SELECT
                'Factura' AS tipo,
                CONCAT('FACT-', LPAD(id::text, 4, '0')) AS numero,
                fecha::text AS fecha,
                total,
                'Completada' AS estado
            FROM ventas

            UNION ALL

            SELECT
                'Orden Compra' AS tipo,
                po AS numero,
                fecha AS fecha,
                COALESCE((
                    SELECT SUM(subtotal)
                    FROM orden_compra_detalle d
                    WHERE d.po = oc.po
                ), 0) AS total,
                estado
            FROM orden_compra oc
        ) AS movimientos
        ORDER BY fecha DESC
        LIMIT 20
    """)

    datos = {
        "productos": execute_scalar("SELECT COUNT(*) FROM productos"),
        "clientes": execute_scalar("SELECT COUNT(*) FROM clientes"),
        "empleados": execute_scalar("SELECT COUNT(*) FROM empleados"),

        "ventas_mes": execute_scalar("""
            SELECT COUNT(*) FROM ventas
            WHERE EXTRACT(MONTH FROM (fecha AT TIME ZONE 'UTC' AT TIME ZONE 'America/Managua')) =
                EXTRACT(MONTH FROM (CURRENT_TIMESTAMP AT TIME ZONE 'America/Managua'))
            AND EXTRACT(YEAR FROM (fecha AT TIME ZONE 'UTC' AT TIME ZONE 'America/Managua')) =
                EXTRACT(YEAR FROM (CURRENT_TIMESTAMP AT TIME ZONE 'America/Managua'))
        """),

        "ingresos_mes": float(execute_scalar("""
            SELECT COALESCE(SUM(total), 0)
            FROM ventas
            WHERE EXTRACT(MONTH FROM (fecha AT TIME ZONE 'UTC' AT TIME ZONE 'America/Managua')) =
                EXTRACT(MONTH FROM (CURRENT_TIMESTAMP AT TIME ZONE 'America/Managua'))
            AND EXTRACT(YEAR FROM (fecha AT TIME ZONE 'UTC' AT TIME ZONE 'America/Managua')) =
                EXTRACT(YEAR FROM (CURRENT_TIMESTAMP AT TIME ZONE 'America/Managua'))
        """) or 0),

        "po_recibidas": execute_scalar("SELECT COUNT(*) FROM orden_compra WHERE estado='Recibida'"),
        "po_pendientes": execute_scalar("SELECT COUNT(*) FROM orden_compra WHERE estado='Pendiente'"),

        "total_po_recibidas": float(execute_scalar("""
            SELECT COALESCE(SUM(d.subtotal), 0)
            FROM orden_compra_detalle d
            INNER JOIN orden_compra o ON d.po=o.po
            WHERE o.estado='Recibida'
        """) or 0),

        "mas_vendidos": mas_vendidos,
        "actividad": actividad,
        "ventas_dias": ventas_7,
        "chart_labels": [r["dia"] for r in ventas_7],
        "chart_values": [float(r["total"] or 0) for r in ventas_7],
    }

    return render_template("dashboard.html", datos=datos)


# =========================
# CLIENTES
# =========================
@app.route("/clientes", methods=["GET", "POST"])
@login_required
def clientes():
    if request.method == "POST":
        execute("INSERT INTO clientes(nombre, cedula, telefono, correo) VALUES (?, ?, ?, ?)", (
            request.form["nombre"], request.form.get("cedula"), request.form.get("telefono"), request.form.get("correo")
        ))
        flash("Cliente guardado correctamente", "success")
        return redirect(url_for("clientes"))
    q = request.args.get("q", "")
    if q:
        like = f"%{q}%"
        rows = fetch_all("""
            SELECT id, nombre, cedula, telefono, correo FROM clientes
            WHERE nombre LIKE ? OR cedula LIKE ? OR telefono LIKE ? OR correo LIKE ?
            ORDER BY id DESC
        """, (like, like, like, like))
    else:
        rows = fetch_all("SELECT id, nombre, cedula, telefono, correo FROM clientes ORDER BY id DESC")
    return render_template("clientes.html", clientes=rows, q=q)


@app.route("/clientes/<int:id>/editar", methods=["POST"])
@login_required
def editar_cliente(id):
    execute("UPDATE clientes SET nombre=?, cedula=?, telefono=?, correo=? WHERE id=?", (
        request.form["nombre"], request.form.get("cedula"), request.form.get("telefono"), request.form.get("correo"), id
    ))
    flash("Cliente actualizado", "success")
    return redirect(url_for("clientes"))


@app.route("/clientes/<int:id>/eliminar", methods=["POST"])
@login_required
def eliminar_cliente(id):
    execute("DELETE FROM clientes WHERE id=?", (id,))
    flash("Cliente eliminado", "success")
    return redirect(url_for("clientes"))


# =========================
# CATEGORÍAS Y PRODUCTOS
# =========================
@app.route("/productos")
@login_required
def productos():
    q = request.args.get("q", "").strip()
    params = []
    sql = """
        SELECT p.id, p.nombre, p.precio, p.stock, p.unidad, p.categoria_id,
               COALESCE(c.nombre, 'Sin categoría') AS categoria,
               (p.precio * p.stock) AS invertido
        FROM productos p
        LEFT JOIN categorias c ON p.categoria_id = c.id
    """
    if q:
        sql += " WHERE p.nombre LIKE ? OR c.nombre LIKE ?"
        params = [f"%{q}%", f"%{q}%"]
    sql += " ORDER BY p.nombre"
    rows = fetch_all(sql, tuple(params))

    resumen = {
        "productos": len(rows),
        "unidades": sum(int(r["stock"] or 0) for r in rows),
        "invertido": sum(float(r["invertido"] or 0) for r in rows),
    }
    return render_template("productos.html", productos=rows, resumen=resumen, q=q)


@app.route("/categorias", methods=["GET", "POST"])
@login_required
@admin_required
def categorias():
    if request.method == "POST":
        accion = request.form.get("accion", "producto")
        if accion == "categoria":
            nombre = request.form.get("nombre_categoria", "").strip()
            if not nombre:
                flash("Ingrese el nombre de la categoría", "warning")
            else:
                execute("INSERT INTO categorias(nombre, descripcion) VALUES (?, ?)", (nombre, ""))
                flash("Categoría agregada correctamente", "success")
            return redirect(url_for("categorias"))

        nombre = request.form.get("nombre", "").strip()
        categoria_id = request.form.get("categoria_id")
        precio = request.form.get("precio") or 0
        stock = request.form.get("stock") or 0
        unidad = request.form.get("unidad", "UND")
        if not nombre or not categoria_id:
            flash("Complete todos los campos del producto", "warning")
        else:
            execute(
                "INSERT INTO productos(nombre, categoria_id, precio, stock, unidad) VALUES (?, ?, ?, ?, ?)",
                (
                    nombre,
                    categoria_id,
                    precio,
                    stock,
                    unidad
                )
            )
            flash("Producto guardado correctamente", "success")
        return redirect(url_for("categorias"))

    q = request.args.get("q", "").strip()
    categorias = fetch_all("SELECT id, nombre FROM categorias ORDER BY nombre")
    params = []
    sql = """
        SELECT p.id,
            p.nombre,
            p.precio,
            p.stock,
            p.unidad,
            p.categoria_id,
            c.nombre AS categoria
        FROM productos p
        INNER JOIN categorias c ON p.categoria_id = c.id
    """
    if q:
        sql += " WHERE p.nombre LIKE ? OR c.nombre LIKE ?"
        params = [f"%{q}%", f"%{q}%"]
    sql += " ORDER BY p.id DESC"
    productos = fetch_all(sql, tuple(params))
    return render_template("categorias.html", productos=productos, categorias=categorias, q=q)


@app.route("/productos/<int:id>/editar", methods=["POST"])
@login_required
@admin_required
def editar_producto(id):
    execute("""
        UPDATE productos
        SET nombre=?,
            categoria_id=?,
            precio=?,
            stock=?,
            unidad=?
        WHERE id=?
    """, (
        request.form["nombre"],
        request.form.get("categoria_id"),
        request.form.get("precio") or 0,
        request.form.get("stock") or 0,
        request.form.get("unidad", "UND"),
        id
    ))
    flash("Producto actualizado correctamente", "success")
    return redirect(request.form.get("next") or url_for("categorias"))


@app.route("/productos/<int:id>/eliminar", methods=["POST"])
@login_required
@admin_required
def eliminar_producto(id):
    execute("DELETE FROM productos WHERE id=?", (id,))
    flash("Producto eliminado correctamente", "success")
    return redirect(request.form.get("next") or url_for("categorias"))



# =========================
# EMPLEADOS
# =========================
@app.route("/empleados", methods=["GET", "POST"])
@login_required
@admin_required
def empleados():
    if request.method == "POST":
        execute("""
            INSERT INTO empleados(nombre, apellido, cedula, telefono, cargo, salario, estado)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (request.form["nombre"], request.form.get("apellido"), request.form.get("cedula"), request.form.get("telefono"), request.form.get("cargo"), request.form.get("salario") or 0, request.form.get("estado", "Activo")))
        flash("Empleado guardado", "success")
        return redirect(url_for("empleados"))
    rows = fetch_all("SELECT id, nombre, apellido, cedula, telefono, cargo, salario, estado FROM empleados ORDER BY id DESC")
    return render_template("empleados.html", empleados=rows)


@app.route("/empleados/<int:id>/eliminar", methods=["POST"])
@login_required
@admin_required
def eliminar_empleado(id):
    execute("DELETE FROM empleados WHERE id=?", (id,))
    flash("Empleado eliminado", "success")
    return redirect(url_for("empleados"))


# =========================
# VENTAS
# =========================
@app.route("/ventas", methods=["GET", "POST"])
@login_required
def ventas():
    if request.method == "POST":
        detalle_json = request.form.get("detalle_json", "[]")
        empleado_id = request.form.get("empleado_id") or None
        cliente_id = request.form.get("cliente_id") or None
        observacion = request.form.get("observacion") or ""

        try:
            detalle = json.loads(detalle_json)
        except Exception:
            detalle = []

        conn = get_connection()
        cursor = conn.cursor()

        try:
            total = 0
            items = []

            for item in detalle:
                producto_id = item.get("producto_id")
                cantidad = int(item.get("cantidad", 0))

                if not producto_id or cantidad <= 0:
                    continue

                cursor.execute(
                    "SELECT id, nombre, precio, stock, unidad FROM productos WHERE id=%s",
                    (producto_id,)
                )
                p = cursor.fetchone()

                if not p:
                    raise Exception("Producto no encontrado")

                if int(p["stock"]) < cantidad:
                    raise Exception(f"Stock insuficiente para {p['nombre']}. Disponible: {p['stock']}")

                subtotal = float(p["precio"]) * cantidad
                total += subtotal

                items.append((p["id"], p["nombre"], cantidad, float(p["precio"]), subtotal))

            if not items:
                raise Exception("Agrega al menos un producto")

            iva = 0
            total_con_iva = total

            cursor.execute("""
                INSERT INTO ventas(fecha, total, empleado_id, cliente_id, observacion)
                VALUES (CURRENT_TIMESTAMP, %s, %s, %s, %s)
                RETURNING id
            """, (total_con_iva, empleado_id, cliente_id, observacion))

            venta_id = cursor.fetchone()["id"]

            for producto_id, nombre, cantidad, precio, subtotal in items:
                cursor.execute("""
                    UPDATE productos
                    SET stock = stock - %s
                    WHERE id = %s AND stock >= %s
                """, (cantidad, producto_id, cantidad))

                if cursor.rowcount == 0:
                    raise Exception(f"No se pudo descontar stock de {nombre}")

                cursor.execute("""
                    INSERT INTO detalle_venta(venta_id, producto_id, cantidad, precio, subtotal)
                    VALUES (%s, %s, %s, %s, %s)
                """, (venta_id, producto_id, cantidad, precio, subtotal))

            conn.commit()
            flash(f"Venta #{venta_id} registrada correctamente", "success")
            return redirect(url_for("factura_pdf", venta_id=venta_id))

        except Exception as e:
            conn.rollback()
            flash(str(e), "danger")

        finally:
            conn.close()

    productos = fetch_all("""
        SELECT id, nombre, precio, stock, unidad
        FROM productos
        WHERE stock > 0
        ORDER BY nombre
    """)

    empleados = fetch_all("""
        SELECT id, nombre, apellido
        FROM empleados
        WHERE estado='Activo'
        ORDER BY nombre
    """)

    clientes = fetch_all("""
        SELECT id, nombre
        FROM clientes
        ORDER BY nombre
    """)

    factura_numero = f"{int(execute_scalar('SELECT COALESCE(MAX(id),0)+1 FROM ventas') or 1):08d}"

    return render_template(
        "ventas.html",
        productos=productos,
        empleados=empleados,
        clientes=clientes,
        factura_numero=factura_numero,
        fecha_hoy=datetime.now().strftime("%d/%m/%Y")
    )

@app.route("/factura/<int:venta_id>.pdf")
@login_required
def factura_pdf(venta_id):
    venta = fetch_one("""
        SELECT 
            v.id,
            v.fecha,
            v.total,
            v.observacion,
            e.nombre AS empleado_nombre,
            e.apellido AS empleado_apellido,
            c.nombre AS cliente_nombre
        FROM ventas v
        LEFT JOIN empleados e ON v.empleado_id = e.id
        LEFT JOIN clientes c ON v.cliente_id = c.id
        WHERE v.id=?
    """, (venta_id,))

    detalles = fetch_all("""
        SELECT 
            p.id AS producto_id,
            p.nombre,
            p.unidad,
            d.cantidad,
            d.precio,
            d.subtotal
        FROM detalle_venta d
        INNER JOIN productos p ON d.producto_id = p.id
        WHERE d.venta_id=?
    """, (venta_id,))

    if not venta:
        flash("Factura no encontrada", "danger")
        return redirect(url_for("ventas"))

    path = os.path.join("facturas_pdf", f"Factura_{venta_id}.pdf")
    c = canvas.Canvas(path, pagesize=letter)

    logo_path = os.path.join("static", "img", "YE.png")
    if os.path.exists(logo_path):
        c.drawImage(logo_path, 40, 705, width=85, height=70, preserveAspectRatio=True, mask="auto")

    c.setFont("Helvetica-Bold", 18)
    c.drawString(135, 755, "FERRETERÍA YAMILON")

    c.setFont("Helvetica", 9)
    c.drawString(135, 740, "Sistema de Facturación e Inventario")
    c.drawString(135, 727, "Rivas, Tola, El Palmar")
    c.drawString(135, 714, "Tel: 7825-6818")

    numero_factura = venta_id

    c.setFont("Helvetica-Bold", 16)
    c.drawRightString(550, 755, "Factura de Venta")

    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(550, 737, f"#FACT-{numero_factura:04d}")

    c.setFont("Helvetica", 9)
    c.drawRightString(550, 722, venta["fecha"].strftime("%d/%m/%Y"))

    cliente = venta["cliente_nombre"] or "CLIENTE GENERAL"

    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, 685, f"CLIENTE: {cliente}")

    y = 650

    c.setFillColorRGB(0.82, 0.82, 0.82)
    c.rect(40, y, 520, 20, fill=1, stroke=0)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(50, y + 7, "Factura")
    c.drawString(200, y + 7, "Fecha")
    c.drawString(380, y + 7, "Vendedor")

    y -= 22

    vendedor = f"{venta['empleado_nombre'] or ''} {venta['empleado_apellido'] or ''}".strip()

    c.setFont("Helvetica", 9)
    c.drawString(50, y + 7, f"FACT-{numero_factura:04d}")
    c.drawString(200, y + 7, venta["fecha"].strftime("%d/%m/%Y"))
    c.drawString(380, y + 7, vendedor or "No registrado")

    y -= 45

    c.setFillColorRGB(0.82, 0.82, 0.82)
    c.rect(40, y, 520, 22, fill=1, stroke=0)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(50, y + 8, "Código")
    c.drawString(105, y + 8, "Descripción")
    c.drawString(310, y + 8, "Cant")
    c.drawString(360, y + 8, "Precio")
    c.drawString(455, y + 8, "Total Neto")

    y -= 24

    c.setFont("Helvetica", 8)
    subtotal_general = 0

    for d in detalles:
        if y < 180:
            c.showPage()
            y = 750

        subtotal_general += float(d["subtotal"] or 0)

        nombre_producto = str(d["nombre"])
        lineas = wrap(nombre_producto, width=38)

        c.drawString(50, y, str(d["producto_id"]).zfill(4))

        y_linea = y
        for linea in lineas:
            c.drawString(105, y_linea, linea)
            y_linea -= 10

        unidad = d["unidad"] or "UND"
        c.drawRightString(330, y, f'{d["cantidad"]} {unidad}')
        c.drawRightString(420, y, f"C${float(d['precio']):,.2f}")
        c.drawRightString(550, y, f"C${float(d['subtotal']):,.2f}")

        y -= max(24, len(lineas) * 12)

    iva = 0
    total = subtotal_general

    # TOTALES: quedan debajo de los productos
    y_totales = y - 20

    c.setFillColorRGB(0.86, 0.86, 0.86)
    c.rect(360, y_totales - 75, 200, 85, fill=1, stroke=0)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(375, y_totales - 5, "Subtotal")
    c.drawRightString(550, y_totales - 5, f"C${subtotal_general:,.2f}")
    c.drawString(375, y_totales - 28, "IVA 0%")
    c.drawRightString(550, y_totales - 28, f"C${iva:,.2f}")
    c.drawString(375, y_totales - 52, "Total a Pagar")
    c.drawRightString(550, y_totales - 52, f"C${total:,.2f}")

    # OBSERVACIONES: abajo y sin mover los totales
    observacion = venta["observacion"] or ""
    if observacion.strip():
        y_obs = 210
        c.setFont("Helvetica-Bold", 9)
        c.drawString(40, y_obs, "OBSERVACIONES:")
        y_obs -= 12
        c.setFont("Helvetica", 8)
        for linea in wrap(observacion, width=85):
            c.drawString(40, y_obs, linea)
            y_obs -= 10

    c.save()
    return send_file(path, as_attachment=False)

@app.route("/gestion-ordenes")
@login_required
@admin_required
def gestion_ordenes():
    q = request.args.get("q", "").strip()
    if q:
        like = f"%{q}%"
        ordenes = fetch_all("""
            SELECT oc.id, oc.po, COALESCE(ir.ir, '') AS ir, oc.proveedor, oc.fecha, oc.estado, oc.memo
            FROM orden_compra oc
            LEFT JOIN ingreso_recepcion ir ON ir.po = oc.po
            WHERE oc.po LIKE ? OR oc.proveedor LIKE ? OR ir.ir LIKE ?
            ORDER BY oc.id DESC
        """, (like, like, like))
    else:
        ordenes = fetch_all("""
            SELECT oc.id, oc.po, COALESCE(ir.ir, '') AS ir, oc.proveedor, oc.fecha, oc.estado, oc.memo
            FROM orden_compra oc
            LEFT JOIN ingreso_recepcion ir ON ir.po = oc.po
            ORDER BY oc.id DESC
        """)
    detalles = {}
    for o in ordenes:
        detalles[o["po"]] = fetch_all("SELECT producto, cantidad, precio, subtotal FROM orden_compra_detalle WHERE po=?", (o["po"],))
    return render_template("gestion_ordenes.html", ordenes=ordenes, detalles=detalles, q=q)


@app.route("/gestion-ordenes/<po>/editar", methods=["GET", "POST"])
@login_required
@admin_required
def editar_orden(po):
    orden = fetch_one("""
        SELECT po, fecha, proveedor, estado, memo
        FROM orden_compra
        WHERE po=?
    """, (po,))

    if not orden:
        flash("Orden no encontrada", "danger")
        return redirect(url_for("gestion_ordenes"))

    if orden["estado"] != "Pendiente":
        flash("Solo se pueden editar órdenes pendientes", "warning")
        return redirect(url_for("gestion_ordenes"))

    if request.method == "POST":
        proveedor = request.form.get("proveedor", "").strip()
        fecha = request.form.get("fecha")
        memo = request.form.get("memo", "")
        productos = request.form.getlist("producto[]")
        cantidades = request.form.getlist("cantidad[]")
        precios = request.form.getlist("precio[]")
        precios_venta = request.form.getlist("precio_venta[]")

        conn = get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                UPDATE orden_compra
                SET proveedor=%s, fecha=%s, memo=%s
                WHERE po=%s
            """, (proveedor, fecha, memo, po))

            cursor.execute("""
                DELETE FROM orden_compra_detalle
                WHERE po=%s
            """, (po,))

            for i in range(len(productos)):
                producto = productos[i]
                cantidad = int(cantidades[i])
                precio = float(precios[i])
                precio_venta = float(precios_venta[i])
                subtotal = cantidad * precio

                cursor.execute("""
                    INSERT INTO orden_compra_detalle(
                        po, producto, cantidad, precio, precio_venta, subtotal
                    )
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (po, producto, cantidad, precio, precio_venta, subtotal))

            conn.commit()
            flash(f"Orden {po} actualizada correctamente", "success")
            return redirect(url_for("gestion_ordenes"))

        except Exception as e:
            conn.rollback()
            flash(str(e), "danger")

        finally:
            conn.close()

    detalles = fetch_all("""
        SELECT producto, cantidad, precio, precio_venta, subtotal
        FROM orden_compra_detalle
        WHERE po=?
    """, (po,))

    return render_template(
        "editar_orden.html",
        orden=orden,
        detalles=detalles
    )

@app.route("/orden-compra", methods=["GET", "POST"])
@login_required
@admin_required
def orden_compra():
    productos = fetch_all("""
        SELECT id, nombre, stock, unidad
        FROM productos
        ORDER BY nombre
    """)

    if request.method == "POST":
        fecha = request.form.get("fecha")
        proveedor = request.form.get("proveedor", "").strip()
        detalle_json = request.form.get("detalle_json", "[]")
        memo = request.form.get("memo", "")

        try:
            detalle = json.loads(detalle_json)
        except Exception:
            detalle = []

        if not proveedor:
            flash("Ingrese el proveedor", "warning")
            return redirect(url_for("orden_compra"))

        if not detalle:
            flash("Agregue al menos un producto", "warning")
            return redirect(url_for("orden_compra"))

        nuevo_id = int(execute_scalar("SELECT COALESCE(MAX(id),0)+1 FROM orden_compra") or 1)
        po = f"PO-PALMAR-{nuevo_id:04d}"

        conn = get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                INSERT INTO orden_compra(po, fecha, proveedor, estado, memo)
                VALUES (%s, %s, %s, %s, %s)
            """, (po, fecha, proveedor, "Pendiente", memo))

            for item in detalle:
                producto = item.get("producto")
                cantidad = int(item.get("cantidad", 0))
                precio = float(item.get("precio", 0))
                precio_venta = float(item.get("precio_venta", 0))
                subtotal = float(item.get("subtotal", cantidad * precio))

                cursor.execute("""
                    INSERT INTO orden_compra_detalle(
                        po, producto, cantidad, precio, precio_venta, subtotal
                    )
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (po, producto, cantidad, precio, precio_venta, subtotal))

            conn.commit()
            flash(f"Orden {po} guardada correctamente", "success")
            return redirect(url_for("gestion_ordenes"))

        except Exception as e:
            conn.rollback()
            flash(str(e), "danger")

        finally:
            conn.close()

    return render_template(
        "orden_compra.html",
        productos=productos,
        fecha_hoy=datetime.now().strftime("%Y-%m-%d")
    )

@app.route("/gestion-ordenes/<po>/imprimir")
@login_required
@admin_required
def orden_pdf(po):
    orden = fetch_one("""
        SELECT oc.po, oc.proveedor, oc.fecha, oc.estado, oc.memo,
               COALESCE(ir.ir, '') AS ir,
               ir.recibido_por, ir.numero_factura, ir.fecha_recepcion
        FROM orden_compra oc
        LEFT JOIN ingreso_recepcion ir ON ir.po = oc.po
        WHERE oc.po=?
    """, (po,))

    detalles = fetch_all("""
        SELECT d.producto,
               d.cantidad,
               d.precio,
               COALESCE(d.precio_venta, 0) AS precio_venta,
               d.subtotal,
               COALESCE(p.unidad, 'UND') AS unidad
        FROM orden_compra_detalle d
        LEFT JOIN productos p ON p.nombre = d.producto
        WHERE d.po=?
    """, (po,))

    if not orden:
        flash("Orden no encontrada", "danger")
        return redirect(url_for("gestion_ordenes"))

    path = os.path.join("ordenes_pdf", f"{po}.pdf")
    c = canvas.Canvas(path, pagesize=letter)

    # LOGO
    logo_path = os.path.join("static", "img", "YE.png")
    if os.path.exists(logo_path):
        c.drawImage(
            logo_path,
            45,
            675,
            width=120,
            height=100,
            preserveAspectRatio=True,
            mask="auto"
        )

    # ENCABEZADO IZQUIERDO

    y = 650
    c.setFont("Helvetica-Bold", 18)
    c.drawString(45, y, "FERRETERÍA YAMILON")
    y -= 28

    c.setFont("Helvetica", 9)
    c.drawString(45, y, "YAMIL ENRIQUE RUIZ CAMACHO")
    y -= 12
    c.drawString(45, y, "Teléfono: 7825-6818")
    y -= 12
    c.drawString(45, y, "Rivas, Tola, El Palmar")

    # ENCABEZADO DERECHO
    c.setFont("Helvetica", 20)
    c.drawRightString(550, 745, "Orden de Compra")

    c.setFont("Helvetica-Bold", 13)
    c.drawRightString(550, 725, f"#{orden['po']}")

    c.setFont("Helvetica", 9)
    c.drawRightString(550, 710, str(orden["fecha"]))

    # TOTAL GRANDE
    subtotal = sum(float(d["subtotal"] or 0) for d in detalles)
    iva = 0
    total = subtotal

    c.setFillColorRGB(0.88, 0.88, 0.88)
    c.rect(340, 620, 210, 65, fill=1, stroke=0)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(350, 665, "TOTAL")

    c.setFont("Helvetica-Bold", 18)
    c.drawRightString(540, 635, f"C${total:,.2f}")

    # MEMO
    c.setFillColorRGB(0.92, 0.92, 0.92)
    c.rect(45, 560, 505, 45, fill=1, stroke=0)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(55, 590, "MEMO")

    c.setFont("Helvetica", 8)
    memo = orden["memo"] or "RELLENO PARA ALMACEN"
    c.drawString(55, 575, str(memo)[:90])

    # DATOS GENERALES
    y = 525
    c.setFillColorRGB(0.82, 0.82, 0.82)
    c.rect(45, y, 505, 18, fill=1, stroke=0)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(55, y + 5, "PO")
    c.drawString(185, y + 5, "Comprador")
    c.drawString(290, y + 5, "Proveedor")
    c.drawString(455, y + 5, "Fecha")

    y -= 20

    c.setFont("Helvetica", 9)
    c.drawString(55, y + 5, str(orden["po"]))
    c.drawString(185, y + 5, session.get("usuario", "Administrador"))

    proveedor = str(orden["proveedor"] or "")
    lineas_proveedor = wrap(proveedor, width=32)
    yy = y + 5
    for linea in lineas_proveedor:
        c.drawString(290, yy, linea)
        yy -= 10

    c.drawString(455, y + 5, str(orden["fecha"]))

    y -= max(40, len(lineas_proveedor) * 12)

    # TABLA PRODUCTOS - SOLO PRECIO DE COMPRA EN PDF
    c.setFillColorRGB(0.82, 0.82, 0.82)
    c.rect(45, y, 505, 20, fill=1, stroke=0)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 7)
    c.drawString(55, y + 6, "Artículo")
    c.drawString(300, y + 6, "Cant.")
    c.drawString(380, y + 6, "Precio")
    c.drawString(470, y + 6, "Valor")

    y -= 24
    c.setFont("Helvetica", 8)

    for d in detalles:
        if y < 140:
            c.showPage()
            y = 750

        unidad = d["unidad"] or "UND"
        producto = str(d["producto"] or "")
        lineas = wrap(producto, width=45)

        yy = y
        for linea in lineas:
            c.drawString(55, yy, linea)
            yy -= 10

        c.drawRightString(340, y, f"{d['cantidad']} {unidad}")
        c.drawRightString(430, y, f"C${float(d['precio'] or 0):,.2f}")
        c.drawRightString(540, y, f"C${float(d['subtotal'] or 0):,.2f}")

        y -= max(20, len(lineas) * 12)

    # RESUMEN
    y -= 15
    if y < 120:
        c.showPage()
        y = 750

    c.setFillColorRGB(0.88, 0.88, 0.88)
    c.rect(360, y - 65, 190, 75, fill=1, stroke=0)

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 9)

    c.drawString(375, y - 5, "Subtotal")
    c.drawRightString(540, y - 5, f"C${subtotal:,.2f}")

    c.drawString(375, y - 27, "IVA 0%")
    c.drawRightString(540, y - 27, f"C${iva:,.2f}")

    c.drawString(375, y - 50, "Total")
    c.drawRightString(540, y - 50, f"C${total:,.2f}")

    # RECIBIDO
    if orden["recibido_por"]:
        c.setFont("Helvetica", 9)
        c.drawString(55, 90, f"Recibido por: {orden['recibido_por']}")
        c.drawString(55, 75, f"Factura proveedor: {orden['numero_factura'] or ''}")

    c.save()
    return send_file(path, as_attachment=False)

@app.route("/gestion-ordenes/<po>/recibir", methods=["POST"])
@login_required
@admin_required
def recibir_orden(po):
    recibido_por = request.form.get("recibido_por", session.get("usuario"))
    numero_factura = request.form.get("numero_factura", "")
    memo = request.form.get("memo", "")

    ir = f"IR-PALMAR-{int(execute_scalar('SELECT COALESCE(MAX(id),0)+1 FROM ingreso_recepcion') or 1):04d}"

    detalles = fetch_all("""
        SELECT producto, cantidad, precio_venta
        FROM orden_compra_detalle
        WHERE po=?
    """, (po,))

    if not detalles:
        flash("Esta orden no tiene productos para recibir", "danger")
        return redirect(url_for("gestion_ordenes"))

    conn = get_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            INSERT INTO ingreso_recepcion(ir, po, recibido_por, numero_factura, memo)
            VALUES (%s, %s, %s, %s, %s)
        """, (ir, po, recibido_por, numero_factura, memo))

        for d in detalles:
            precio_venta = float(d["precio_venta"] or 0)

            if precio_venta <= 0:
                raise Exception(f"El producto {d['producto']} no tiene precio de venta válido")

            cursor.execute("""
                UPDATE productos
                SET stock = stock + %s,
                    precio = %s
                WHERE nombre = %s
            """, (d["cantidad"], precio_venta, d["producto"]))

            if cursor.rowcount == 0:
                raise Exception(f"No se encontró el producto en inventario: {d['producto']}")

        cursor.execute("""
            UPDATE orden_compra
            SET estado='Recibida'
            WHERE po=%s
        """, (po,))

        conn.commit()
        flash(f"Orden {po} recibida correctamente. IR generado: {ir}", "success")

    except Exception as e:
        conn.rollback()
        flash(str(e), "danger")

    finally:
        conn.close()

    return redirect(url_for("gestion_ordenes"))

@app.route("/gestion-ordenes/<po>/cerrar", methods=["POST"])
@login_required
@admin_required
def cerrar_orden(po):
    try:
        execute("""
            UPDATE orden_compra
            SET estado='Cerrada'
            WHERE po=?
        """, (po,))

        flash(f"Orden {po} cerrada correctamente", "success")

    except Exception as e:
        flash(str(e), "danger")

    return redirect(url_for("gestion_ordenes"))

@app.route("/facturas")
@login_required
def facturas():
    ventas = fetch_all("""
        SELECT id, fecha, total
        FROM ventas
        ORDER BY id DESC
    """)

    return render_template("facturas.html", ventas=ventas)

@app.route("/inventario/excel")
@login_required
def descargar_inventario_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    productos = fetch_all("""
        SELECT
            p.nombre AS producto,
            COALESCE(c.nombre, 'Sin categoría') AS categoria,
            p.precio,
            p.stock,
            p.unidad,
            (p.precio * p.stock) AS total
        FROM productos p
        LEFT JOIN categorias c
            ON p.categoria_id = c.id
        ORDER BY p.nombre
    """)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    encabezados = ["Producto", "Categoría", "Precio", "Stock", "Unidad", "Total"]
    ws.append(encabezados)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111111")
        cell.alignment = Alignment(horizontal="center")

    for p in productos:
        ws.append([
            p["producto"],
            p["categoria"],
            float(p["precio"] or 0),
            int(p["stock"] or 0),
            p["unidad"] or "UND",
            float(p["total"] or 0)
        ])

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 15

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = 'C$#,##0.00'

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = 'C$#,##0.00'

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="Inventario_Palmar.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
